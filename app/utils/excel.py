import io
import json
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

ILLER_DOSYASI = Path(__file__).resolve().parent.parent / "data" / "iller.js"
MUSTERI_SUTUNLARI = ["Firma Adı", "Yetkili", "Telefon", "E-Posta", "Vergi Dairesi", "Vergi No", "İl", "İlçe", "Müşteri Türü", "Adres", "Açıklama", "Durum"]
TURLER = ["Alıcı", "Tedarikçi", "Satıcı"]
URUN_SUTUNLARI = ["Ürün Kodu", "Ürün Adı", "Marka", "Model", "Ürün Türü", "Ürün Sınıfı", "Ürün Cinsi", "İstasyon Kodları", "Birim", "Mevcut Stok", "Min. Stok", "Max. Stok", "Maliyet", "Satış Fiyatı", "Açıklama", "Durum"]
ESKI_URUN_SUTUNLARI = [sutun for sutun in URUN_SUTUNLARI if sutun != "İstasyon Kodları"]
PERSONEL_SUTUNLARI = ["Ad Soyad", "İstasyon Kodları", "Görev", "Durum"]
ISTASYON_SUTUNLARI = ["İstasyon Kodu", "İstasyon Adı", "Bölüm", "Açıklama", "Durum"]
MAKINE_SUTUNLARI = ["Makine Kodu", "Makine Adı", "İstasyon Kodu", "Model", "Kapasite", "Durum"]


def il_ilce_verisi():
    with ILLER_DOSYASI.open(encoding="utf-8") as dosya:
        return json.load(dosya)


def metin(deger):
    return str(deger or "").strip()


def sayi(deger, alan):
    if deger in (None, ""):
        return 0.0
    try:
        return float(deger)
    except (TypeError, ValueError):
        raise ValueError(f"{alan} sayısal olmalı")


def sonraki_musteri_kodu(db: Session, kullanilan_kodlar: set[str]) -> str:
    """İçe aktarım süresince tekrar etmeyen bir müşteri kodu üretir."""
    en_yuksek_numara = 0
    for kod in kullanilan_kodlar:
        if kod and kod.startswith("M") and kod[1:].isdigit():
            en_yuksek_numara = max(en_yuksek_numara, int(kod[1:]))

    sira = en_yuksek_numara + 1
    kod = f"M{sira:06}"
    while kod in kullanilan_kodlar:
        sira += 1
        kod = f"M{sira:06}"
    kullanilan_kodlar.add(kod)
    return kod


def sonraki_personel_kodu(kullanilan_kodlar: set[str]) -> str:
    """Excel aktarımında yeni personel için P000001 biçiminde kod üretir."""
    en_yuksek_numara = 0
    for kod in kullanilan_kodlar:
        if kod and kod.startswith("P") and kod[1:].isdigit():
            en_yuksek_numara = max(en_yuksek_numara, int(kod[1:]))
    sira = en_yuksek_numara + 1
    kod = f"P{sira:06}"
    while kod in kullanilan_kodlar:
        sira += 1
        kod = f"P{sira:06}"
    kullanilan_kodlar.add(kod)
    return kod


def excel_satirlarini_oku(dosya_icerigi):
    # Excel masaüstü uygulaması bazı açılır liste doğrulamalarını OpenPyXL'in
    # desteklemediği x14 uzantısına dönüştürebilir. Bu uyarı hücre verisini
    # etkilemez; yalnızca doğrulama tanımını kaldırır ve aktarımı durdurmamalıdır.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"Data Validation extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\\.worksheet\\._reader",
        )
        kitap = openpyxl.load_workbook(io.BytesIO(dosya_icerigi), data_only=True)
    referans_zamani = ""
    if "Sistem Bilgileri" in kitap.sheetnames:
        for satir in kitap["Sistem Bilgileri"].iter_rows(min_col=1, max_col=2, values_only=True):
            if metin(satir[0]) == "Aktarım Referansı (UTC)":
                referans_zamani = metin(satir[1])
                break
    if "Müşteriler" not in kitap.sheetnames:
        return [], [], ["'Müşteriler' sayfası bulunamadı."]
    sayfa = kitap["Müşteriler"]
    basliklar = [metin(hucre.value) for hucre in sayfa[1]]
    if basliklar[:len(MUSTERI_SUTUNLARI)] != MUSTERI_SUTUNLARI:
        return [], [], ["Müşteriler sayfasındaki sütun başlıkları taslakla uyuşmuyor."]

    iller = il_ilce_verisi()
    satirlar, hatalar = [], []
    for sira, satir in enumerate(sayfa.iter_rows(min_row=2, values_only=True), start=2):
        if not any(satir):
            continue
        veri = dict(zip(MUSTERI_SUTUNLARI, map(metin, satir)))
        satir_hatalari = []
        if not veri["Firma Adı"]:
            satir_hatalari.append("Firma adı zorunlu")
        if veri["Müşteri Türü"] not in TURLER:
            satir_hatalari.append("Müşteri türü Alıcı, Tedarikçi veya Satıcı olmalı")
        if veri["Durum"] not in ("Aktif", "Pasif"):
            satir_hatalari.append("Durum Aktif veya Pasif olmalı")
        if veri["İl"] and veri["İl"] not in iller:
            satir_hatalari.append("Geçersiz il")
        if veri["İlçe"] and (not veri["İl"] or veri["İlçe"] not in iller.get(veri["İl"], [])):
            satir_hatalari.append("İlçenin seçilen ille eşleşmesi gerekiyor")
        if satir_hatalari:
            hatalar.append(f"Satır {sira}: {', '.join(satir_hatalari)}")
        else:
            veri["_excel_referans_zamani"] = referans_zamani
            satirlar.append(veri)
    urunler = []
    urun_sayfa_adi = next((ad for ad in ("Hammadde ve Ticari Ürünler", "Stok Ürünleri") if ad in kitap.sheetnames), None)
    if urun_sayfa_adi:
        stok_sayfasi = kitap[urun_sayfa_adi]
        stok_basliklari = [metin(hucre.value) for hucre in stok_sayfasi[1]]
        urun_sutunlari = URUN_SUTUNLARI if stok_basliklari[:len(URUN_SUTUNLARI)] == URUN_SUTUNLARI else ESKI_URUN_SUTUNLARI
        if stok_basliklari[:len(urun_sutunlari)] != urun_sutunlari:
            hatalar.append(f"{urun_sayfa_adi} sayfasındaki sütun başlıkları taslakla uyuşmuyor.")
        else:
            for sira, satir in enumerate(stok_sayfasi.iter_rows(min_row=2, values_only=True), start=2):
                if not any(satir):
                    continue
                veri = dict(zip(urun_sutunlari, satir))
                veri.setdefault("İstasyon Kodları", "")
                satir_hatalari = []
                kod, ad = metin(veri["Ürün Kodu"]), metin(veri["Ürün Adı"])
                urun_turu, birim = metin(veri["Ürün Türü"]), metin(veri["Birim"])
                if not kod:
                    satir_hatalari.append("Ürün kodu zorunlu")
                if not ad:
                    satir_hatalari.append("Ürün adı zorunlu")
                if not urun_turu:
                    satir_hatalari.append("Ürün türü zorunlu")
                if metin(veri["Durum"]) not in ("Aktif", "Pasif"):
                    satir_hatalari.append("Durum Aktif veya Pasif olmalı")
                try:
                    urun = {
                    "kodu": kod, "adi": ad, "marka": metin(veri["Marka"]), "model": metin(veri["Model"]), "urun_tipi": "Hammadde",
                        "stok_turu_adi": urun_turu, "urun_sinifi_anahtari": metin(veri["Ürün Sınıfı"]),
                        "urun_cinsi": metin(veri["Ürün Cinsi"]),
                        "istasyon_kodlari": metin(veri["İstasyon Kodları"]),
                        "birim": birim or "Adet",
                        "mevcut_stok": sayi(veri["Mevcut Stok"], "Mevcut stok"),
                        "min_stok": sayi(veri["Min. Stok"], "Min. stok"),
                        "max_stok": sayi(veri["Max. Stok"], "Max. stok"),
                        "maliyet": sayi(veri["Maliyet"], "Maliyet"),
                        "satis_fiyati": sayi(veri["Satış Fiyatı"], "Satış fiyatı"),
                        "aciklama": metin(veri["Açıklama"]),
                        "aktif": metin(veri["Durum"]) == "Aktif",
                        "_excel_referans_zamani": referans_zamani,
                    }
                except ValueError as hata:
                    satir_hatalari.append(str(hata))
                    urun = None
                if satir_hatalari:
                    hatalar.append(f"{urun_sayfa_adi} satır {sira}: {', '.join(satir_hatalari)}")
                elif urun:
                    urunler.append(urun)
    def liste_sayfasi_oku(sayfa_adlari, sutunlar, etiket, zorunlu_alanlar=None):
        sayfa_adi = next((ad for ad in sayfa_adlari if ad in kitap.sheetnames), None)
        if not sayfa_adi:
            return []
        liste_sayfasi = kitap[sayfa_adi]
        basliklar = [metin(hucre.value) for hucre in liste_sayfasi[1]]
        if basliklar[:len(sutunlar)] != sutunlar:
            hatalar.append(f"{sayfa_adi} sayfasındaki sütun başlıkları taslakla uyuşmuyor.")
            return []
        sonuc = []
        for sira, excel_satiri in enumerate(liste_sayfasi.iter_rows(min_row=2, values_only=True), start=2):
            if not any(excel_satiri):
                continue
            veri = dict(zip(sutunlar, excel_satiri))
            satir_hatalari = []
            zorunlu_alanlar = zorunlu_alanlar or sutunlar[:2]
            eksik_alanlar = [alan for alan in zorunlu_alanlar if not metin(veri[alan])]
            if eksik_alanlar:
                satir_hatalari.append(f"{', '.join(eksik_alanlar)} zorunlu")
            if metin(veri["Durum"]) not in ("Aktif", "Pasif"):
                satir_hatalari.append("Durum Aktif veya Pasif olmalı")
            if satir_hatalari:
                hatalar.append(f"{etiket} satır {sira}: {', '.join(satir_hatalari)}")
            else:
                veri["_excel_referans_zamani"] = referans_zamani
                sonuc.append(veri)
        return sonuc

    personeller = liste_sayfasi_oku(
        ("Personel Listesi", "Personeller", "Çalışanlar"), PERSONEL_SUTUNLARI, "Personeller", ["Ad Soyad"]
    )
    istasyonlar = liste_sayfasi_oku(
        ("İstasyon Listesi", "İstasyonlar", "İstasyon"), ISTASYON_SUTUNLARI, "İstasyonlar"
    )
    makineler = liste_sayfasi_oku(
        ("Makine Listesi", "Makineler", "Makine"), MAKINE_SUTUNLARI, "Makineler"
    )
    for sira, makine in enumerate(makineler, start=2):
        if not metin(makine["İstasyon Kodu"]):
            hatalar.append(f"Makineler satır {sira}: İstasyon Kodu zorunlu")
    return satirlar, urunler, personeller, istasyonlar, makineler, hatalar

def excel_sablonu_olustur(sablon_verisi) -> bytes:
    iller = il_ilce_verisi()
    firma = sablon_verisi["firma"]
    mevcut_musteriler = sablon_verisi["musteriler"]
    mevcut_urunler = sablon_verisi["urunler"]
    mevcut_personeller = sablon_verisi["personeller"]
    mevcut_istasyonlar = sablon_verisi["istasyonlar"]
    mevcut_makineler = sablon_verisi["makineler"]
    mevcut_stok_turleri = sablon_verisi.get("stok_turleri", [])
    mevcut_urun_siniflari = sablon_verisi.get("siniflar", [])
    urun_istasyon_kodlari = sablon_verisi.get("urun_istasyon_kodlari", {})
    kitap = openpyxl.Workbook()
    sistem = kitap.active
    sistem.title = "Sistem Bilgileri"
    sistem.append(["MÜYS Excel Aktarım ve Dışa Aktarma"])
    sistem.append(["Açıklama", "Firma, müşteri, stok ve üretim ana verileri ayrı sayfalarda yer alır."])
    sistem.append(["Kurallar", "Personel ve hammadde sayfalarında birden fazla istasyon kodunu virgülle ayırın: KESIM, BUKUM."])
    sistem.append(["Dışa Aktarım Tarihi", datetime.now().strftime("%d.%m.%Y %H:%M")])
    sistem.append(["Aktarım Referansı (UTC)", datetime.utcnow().isoformat(timespec="microseconds")])
    sistem.append(["Sistem", "MÜYS v0.1.1 / FastAPI / SQLite"])
    sistem.append(["Kayıt Özeti", f"{len(mevcut_musteriler)} müşteri, {len(mevcut_urunler)} stok ürünü, {len(mevcut_personeller)} personel, {len(mevcut_istasyonlar)} istasyon, {len(mevcut_makineler)} makine"])
    sistem.append([])
    sistem.append(["Firma Bilgileri"])
    for etiket, alan in [("Firma Adı", "firma_adi"), ("Vergi No", "vergi_no"), ("Vergi Dairesi", "vergi_dairesi"), ("Telefon", "telefon"), ("E-Posta", "email"), ("Web Sitesi", "web_sitesi"), ("Adres", "adres")]:
        sistem.append([etiket, getattr(firma, alan, "") if firma else ""])
    sistem.column_dimensions["A"].width = 28
    sistem.column_dimensions["B"].width = 100
    for hucre in sistem[1]:
        hucre.font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")

    musteriler = kitap.create_sheet("Müşteriler")
    musteriler.append(MUSTERI_SUTUNLARI)
    for musteri in mevcut_musteriler:
        musteriler.append([
            musteri.firma_adi, musteri.yetkili, musteri.telefon, musteri.email,
            musteri.vergi_dairesi, musteri.vergi_no, musteri.il, musteri.ilce,
            musteri.musteri_turu or "Alıcı", musteri.adres, musteri.aciklama,
            "Aktif" if musteri.aktif else "Pasif",
        ])
    musteriler.freeze_panes = "A2"
    son_satir = max(501, len(mevcut_musteriler) + 1)
    musteriler.auto_filter.ref = f"A1:L{son_satir}"
    for hucre in musteriler[1]:
        hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    for sutun, genislik in zip("ABCDEFGHIJKL", [28, 22, 18, 28, 20, 16, 18, 20, 18, 45, 35, 14]):
        musteriler.column_dimensions[sutun].width = genislik

    istasyon_kodlari = {istasyon.id: istasyon.kodu for istasyon in mevcut_istasyonlar}
    personel_istasyon_kodlari = sablon_verisi.get("personel_istasyon_kodlari", {})
    for sayfa_adi, basliklar, satirlar, genislikler in [
        ("Personel Listesi", PERSONEL_SUTUNLARI, [[p.ad_soyad, ", ".join(personel_istasyon_kodlari.get(p.id, [])), p.gorev, "Aktif" if p.aktif else "Pasif"] for p in mevcut_personeller], [30, 28, 22, 14]),
        ("İstasyon Listesi", ["İstasyon Kodu", "İstasyon Adı", "Bölüm", "Açıklama", "Durum"], [[i.kodu, i.adi, i.bolum, i.aciklama, "Aktif" if i.aktif else "Pasif"] for i in mevcut_istasyonlar], [18, 30, 22, 36, 14]),
        ("Makine Listesi", ["Makine Kodu", "Makine Adı", "İstasyon Kodu", "Model", "Kapasite", "Durum"], [[m.kodu, m.adi, istasyon_kodlari.get(m.istasyon_id, ""), m.model, m.kapasite, "Aktif" if m.aktif else "Pasif"] for m in mevcut_makineler], [18, 30, 18, 22, 18, 14]),
    ]:
        sayfa = kitap.create_sheet(sayfa_adi)
        sayfa.append(basliklar)
        for satir in satirlar:
            sayfa.append(satir)
        sayfa.freeze_panes = "A2"
        sayfa.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(basliklar))}{max(2, len(satirlar) + 1)}"
        for hucre in sayfa[1]:
            hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
        for index, genislik in enumerate(genislikler, start=1):
            sayfa.column_dimensions[openpyxl.utils.get_column_letter(index)].width = genislik

    stok_tur_adlari = {tur.id: tur.adi for tur in mevcut_stok_turleri}
    urun_sinif_adlari = {sinif.id: f"{sinif.kodu} · {sinif.adi}" for sinif in mevcut_urun_siniflari}
    aktarilacak_urunler = [u for u in mevcut_urunler if u.urun_tipi in ("Hammadde", "TicariMamul")]
    stok = kitap.create_sheet("Hammadde ve Ticari Ürünler")
    stok.append(URUN_SUTUNLARI)
    for urun in aktarilacak_urunler:
        stok.append([
            urun.kodu, urun.adi, urun.marka or "", urun.model or "", stok_tur_adlari.get(urun.stok_urun_turu_id, "Hammadde"),
            urun_sinif_adlari.get(urun.urun_sinifi_id, ""), urun.urun_cinsi or "",
            ", ".join(urun_istasyon_kodlari.get(urun.id, [])), urun.birim or "Adet",
            urun.mevcut_stok or 0, urun.min_stok or 0, urun.max_stok or 0,
            urun.maliyet or 0, urun.satis_fiyati or 0, urun.aciklama or "",
            "Aktif" if urun.aktif else "Pasif",
        ])
    stok.freeze_panes = "A2"
    stok_son_satir = max(501, len(aktarilacak_urunler) + 1)
    stok.auto_filter.ref = f"A1:P{stok_son_satir}"
    for hucre in stok[1]:
        hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    for sutun, genislik in zip("ABCDEFGHIJKLMNOP", [18, 32, 20, 20, 20, 22, 22, 24, 14, 16, 16, 16, 14, 16, 36, 12]):
        stok.column_dimensions[sutun].width = genislik

    listeler = kitap.create_sheet("Listeler")
    listeler.append(["İller", "İlçeler", "Müşteri Türleri", "Ürün Türleri", "Ürün Sınıfları", "Durumlar", "İstasyon Kodları"])
    tum_ilceler = sorted({ilce for ilceler in iller.values() for ilce in ilceler})
    durumlar = ["Aktif", "Pasif"]
    tur_listesi = [tur.adi for tur in mevcut_stok_turleri]
    sinif_listesi = [f"{sinif.kodu} · {sinif.adi}" for sinif in mevcut_urun_siniflari if sinif.aktif]
    istasyon_listesi = [istasyon.kodu for istasyon in mevcut_istasyonlar]
    for sira in range(max(len(iller), len(tum_ilceler), len(TURLER), len(tur_listesi), len(sinif_listesi), len(durumlar), len(istasyon_listesi))):
        listeler.append([
            sorted(iller)[sira] if sira < len(iller) else None,
            tum_ilceler[sira] if sira < len(tum_ilceler) else None,
            TURLER[sira] if sira < len(TURLER) else None,
            tur_listesi[sira] if sira < len(tur_listesi) else None,
            sinif_listesi[sira] if sira < len(sinif_listesi) else None,
            durumlar[sira] if sira < len(durumlar) else None,
            istasyon_listesi[sira] if sira < len(istasyon_listesi) else None,
        ])
    listeler.sheet_state = "hidden"
    for formül, alan in [("'Listeler'!$A$2:$A$82", f"G2:G{son_satir}"), (f"'Listeler'!$B$2:$B${len(tum_ilceler) + 1}", f"H2:H{son_satir}"), ("'Listeler'!$C$2:$C$4", f"I2:I{son_satir}")]:
        dogrulama = DataValidation(type="list", formula1=formül, allow_blank=True)
        musteriler.add_data_validation(dogrulama)
        dogrulama.add(alan)

    for formül, alan, bos_birakilabilir in [
        (f"'Listeler'!$D$2:$D${max(2, len(tur_listesi) + 1)}", f"E2:E{stok_son_satir}", False),
        (f"'Listeler'!$E$2:$E${max(2, len(sinif_listesi) + 1)}", f"F2:F{stok_son_satir}", True),
        ("'Listeler'!$F$2:$F$3", f"P2:P{stok_son_satir}", False),
    ]:
        dogrulama = DataValidation(type="list", formula1=formül, allow_blank=bos_birakilabilir)
        stok.add_data_validation(dogrulama)
        dogrulama.add(alan)
    musteri_durumu = DataValidation(type="list", formula1="'Listeler'!$F$2:$F$3", allow_blank=False)
    musteriler.add_data_validation(musteri_durumu)
    musteri_durumu.add(f"L2:L{son_satir}")

    for sayfa_adi, alan in [("Personel Listesi", "D2:D1000"), ("İstasyon Listesi", "E2:E1000"), ("Makine Listesi", "F2:F1000")]:
        dogrulama = DataValidation(type="list", formula1="'Listeler'!$F$2:$F$3", allow_blank=False)
        kitap[sayfa_adi].add_data_validation(dogrulama)
        dogrulama.add(alan)
    istasyon_dogrulamasi = DataValidation(type="list", formula1=f"'Listeler'!$G$2:$G${max(2, len(istasyon_listesi) + 1)}", allow_blank=True)
    kitap["Personel Listesi"].add_data_validation(istasyon_dogrulamasi)
    istasyon_dogrulamasi.add("B2:B1000")
    urun_istasyon_dogrulamasi = DataValidation(type="list", formula1=f"'Listeler'!$G$2:$G${max(2, len(istasyon_listesi) + 1)}", allow_blank=True)
    stok.add_data_validation(urun_istasyon_dogrulamasi)
    urun_istasyon_dogrulamasi.add(f"H2:H{stok_son_satir}")

    # Gizli yardımcı liste sayfasından sonra da olsa ürün sayfasını çalışma
    # kitabındaki mutlak son sekme yap.
    kitap.move_sheet(stok, offset=len(kitap.sheetnames))

    akis = io.BytesIO()
    kitap.save(akis)
    return akis.getvalue()
