import io
from datetime import datetime

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

SAYFALAR = {
    "Personeller": ["Personel Kodu", "Ad Soyad", "Departman", "Görev", "Durum"],
    "İstasyonlar": ["İstasyon Kodu", "İstasyon Adı", "Bölüm", "Açıklama", "Durum"],
    "Makineler": ["Makine Kodu", "Makine Adı", "İstasyon Kodu", "Model", "Kapasite", "Durum"],
    "Personel Makine Atamaları": ["Personel Kodu", "Makine Kodu", "Rol", "Hedef Performans", "Durum"],
    "Ürün Sınıfları": ["Sınıf Kodu", "Sınıf Adı", "Açıklama", "Durum"],
    "Sınıf Reçete Operasyonları": ["Sınıf Kodu", "Sıra", "İstasyon Kodu", "Makine Kodu", "Operasyon Adı", "Kontrol Noktası", "Durum"],
    "Ürünler": ["Ürün Kodu", "Ürün Adı", "Ürün Türü", "Ürün Sınıfı Kodu", "Birim", "Mevcut Stok", "Min. Stok", "Max. Stok", "Maliyet", "Satış Fiyatı", "Açıklama", "Durum"],
    "Ürün Reçetesi": ["Üst Ürün Kodu", "Bileşen Ürün Kodu", "Miktar", "Birim", "Fire Oranı (%)", "Sıra", "Hedef Çevrim Süresi (dk)"],
}


def _metin(deger): return str(deger or "").strip()


def excel_dosyasini_oku(icerik: bytes) -> dict[str, list[dict]]:
    kitap = openpyxl.load_workbook(io.BytesIO(icerik), data_only=True)
    veriler = {}
    for ad, sutunlar in SAYFALAR.items():
        if ad not in kitap.sheetnames: raise ValueError(f"'{ad}' sayfası bulunamadı")
        sayfa = kitap[ad]
        basliklar = [_metin(h.value) for h in sayfa[1]]
        if basliklar[:len(sutunlar)] != sutunlar: raise ValueError(f"'{ad}' sütunları şablonla uyuşmuyor")
        veriler[ad] = [dict(zip(sutunlar, satir)) for satir in sayfa.iter_rows(min_row=2, values_only=True) if any(satir)]
    return veriler


def excel_sablonu_olustur(veriler: dict[str, list]) -> bytes:
    kitap = openpyxl.Workbook()
    bilgi = kitap.active
    bilgi.title = "Sistem Bilgileri"
    bilgi.append(["MÜYS Üretim Ana Veri Aktarımı"])
    bilgi.append(["Açıklama", "Sayfaları sırayla doldurun: istasyonlar → makineler → personel/makine atamaları → ürün sınıfları → operasyonlar."])
    bilgi.append(["Tarih", datetime.now().strftime("%d.%m.%Y %H:%M")])
    bilgi.column_dimensions["A"].width, bilgi.column_dimensions["B"].width = 28, 115
    for ad, satirlar in veriler.items():
        sayfa = kitap.create_sheet(ad); sayfa.append(SAYFALAR[ad])
        for satir in satirlar: sayfa.append(satir)
        sayfa.freeze_panes = "A2"
        sayfa.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(SAYFALAR[ad]))}{max(2, len(satirlar) + 1)}"
        for hucre in sayfa[1]:
            hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
        for sutun in range(1, len(SAYFALAR[ad]) + 1): sayfa.column_dimensions[openpyxl.utils.get_column_letter(sutun)].width = 22
    listeler = kitap.create_sheet("Listeler")
    listeler.append(["Durumlar", "Ürün Türleri"])
    for sira in range(4): listeler.append([["Aktif", "Pasif"][sira] if sira < 2 else None, ["Hammadde", "YariMamul", "Mamul", "TicariMamul"][sira]])
    listeler.sheet_state = "hidden"
    for ad, sutun in [("Personeller", "E"), ("İstasyonlar", "E"), ("Makineler", "F"), ("Personel Makine Atamaları", "E"), ("Ürün Sınıfları", "D"), ("Sınıf Reçete Operasyonları", "G"), ("Ürünler", "L")]:
        dogrulama = DataValidation(type="list", formula1="'Listeler'!$A$2:$A$3"); kitap[ad].add_data_validation(dogrulama); dogrulama.add(f"{sutun}2:{sutun}1000")
    dogrulama = DataValidation(type="list", formula1="'Listeler'!$B$2:$B$5"); kitap["Ürünler"].add_data_validation(dogrulama); dogrulama.add("C2:C1000")
    akis = io.BytesIO(); kitap.save(akis); return akis.getvalue()
