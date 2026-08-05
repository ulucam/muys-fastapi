import io

import openpyxl
from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.context import template_data
from app.database import get_db
from app.models.recete import Recete
from app.models.stok_hareket import StokHareket
from app.models.urun import Urun
from app.roles import STOK
from app.security import yetki_kontrol
from app.services.islem_log_service import islem_logla

router = APIRouter(tags=["Stok"])
templates = Jinja2Templates(directory="app/templates")
HAMMADDE_SUTUNLARI = ["Ürün Kodu", "Ürün Adı", "Birim", "Giriş Miktarı", "Min. Stok", "Max. Stok", "Birim Maliyet", "Açıklama", "Durum"]


def metin(deger):
    return str(deger or "").strip()


def sayi(deger, alan, satir):
    try:
        sonuc = float(deger or 0)
    except (TypeError, ValueError):
        raise ValueError(f"Satır {satir}: {alan} sayısal olmalı")
    if sonuc < 0:
        raise ValueError(f"Satır {satir}: {alan} negatif olamaz")
    return sonuc


def urun_sayfasi_verisi(request, db, **ek):
    data = template_data(request)
    data["urunler"] = db.query(Urun).filter(Urun.urun_tipi == "Hammadde").order_by(Urun.kodu).all()
    data.update(ek)
    return data


@router.get("/urunler", response_class=HTMLResponse)
def urunler(request: Request, db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(STOK))):
    return templates.TemplateResponse("stok/urunler.html", urun_sayfasi_verisi(request, db))


@router.get("/urunler/hammadde-sablon")
def hammadde_sablon(yetki=Depends(yetki_kontrol(STOK))):
    kitap = openpyxl.Workbook()
    sayfa = kitap.active
    sayfa.title = "Hammadde Girişi"
    sayfa.append(HAMMADDE_SUTUNLARI)
    sayfa.freeze_panes = "A2"
    sayfa.auto_filter.ref = "A1:I1000"
    genislikler = [18, 34, 14, 18, 15, 15, 18, 42, 14]
    for index, genislik in enumerate(genislikler, 1):
        sayfa.column_dimensions[openpyxl.utils.get_column_letter(index)].width = genislik
    for hucre in sayfa[1]:
        hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    akis = io.BytesIO()
    kitap.save(akis)
    return Response(akis.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=muys-hammadde-giris-sablonu.xlsx"})


@router.post("/urunler/hammadde-aktar", response_class=HTMLResponse)
async def hammadde_aktar(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(STOK))):
    dosya_adi = file.filename or "dosya.xlsx"
    hatalar = []
    if not dosya_adi.lower().endswith(".xlsx"):
        hatalar.append("Yalnızca .xlsx uzantılı dosya yüklenebilir.")
    veriler = {}
    if not hatalar:
        try:
            kitap = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
            if "Hammadde Girişi" not in kitap.sheetnames:
                raise ValueError("'Hammadde Girişi' sayfası bulunamadı")
            sayfa = kitap["Hammadde Girişi"]
            basliklar = [metin(h.value) for h in sayfa[1]]
            if basliklar[:len(HAMMADDE_SUTUNLARI)] != HAMMADDE_SUTUNLARI:
                raise ValueError("Sütun başlıkları hammadde şablonuyla uyuşmuyor")
            for satir_no, satir in enumerate(sayfa.iter_rows(min_row=2, values_only=True), 2):
                if not any(satir):
                    continue
                veri = dict(zip(HAMMADDE_SUTUNLARI, satir))
                kod, ad = metin(veri["Ürün Kodu"]), metin(veri["Ürün Adı"])
                if not kod or not ad:
                    raise ValueError(f"Satır {satir_no}: Ürün Kodu ve Ürün Adı zorunlu")
                if metin(veri["Durum"]) not in ("Aktif", "Pasif"):
                    raise ValueError(f"Satır {satir_no}: Durum Aktif veya Pasif olmalı")
                miktar = sayi(veri["Giriş Miktarı"], "Giriş Miktarı", satir_no)
                kayit = {
                    "kodu": kod, "adi": ad, "birim": metin(veri["Birim"]) or "Kg",
                    "miktar": miktar, "min_stok": sayi(veri["Min. Stok"], "Min. Stok", satir_no),
                    "max_stok": sayi(veri["Max. Stok"], "Max. Stok", satir_no),
                    "maliyet": sayi(veri["Birim Maliyet"], "Birim Maliyet", satir_no),
                    "aciklama": metin(veri["Açıklama"]), "aktif": metin(veri["Durum"]) == "Aktif",
                }
                if kod in veriler:
                    kayit["miktar"] += veriler[kod]["miktar"]
                veriler[kod] = kayit
        except Exception as hata:
            hatalar.append(str(hata))
    if not hatalar and not veriler:
        hatalar.append("Excel dosyasında aktarılacak hammadde satırı bulunamadı.")
    if hatalar:
        return templates.TemplateResponse("stok/urunler.html", urun_sayfasi_verisi(request, db, hatalar=hatalar), status_code=400)
    try:
        urun_haritasi = {u.kodu: u for u in db.query(Urun).all()}
        hareketler = []
        for kod, veri in veriler.items():
            urun = urun_haritasi.get(kod)
            if not urun:
                urun = Urun(kodu=kod, mevcut_stok=0)
                db.add(urun)
                urun_haritasi[kod] = urun
            urun.adi, urun.urun_tipi, urun.birim = veri["adi"], "Hammadde", veri["birim"]
            urun.min_stok, urun.max_stok, urun.maliyet = veri["min_stok"], veri["max_stok"], veri["maliyet"]
            urun.aciklama, urun.aktif = veri["aciklama"], veri["aktif"]
            urun.mevcut_stok = float(urun.mevcut_stok or 0) + veri["miktar"]
            hareketler.append((urun, veri["miktar"]))
        db.flush()
        for urun, miktar in hareketler:
            if miktar:
                db.add(StokHareket(urun_id=urun.id, hareket_tipi="Giriş", miktar=miktar, aciklama="Excel toplu hammadde girişi", referans=dosya_adi[:50]))
        islem_logla(db, request, "Stok", "Toplu hammadde girişi tamamlandı", f"Dosya: {dosya_adi}. {len(veriler)} hammadde işlendi.")
        db.commit()
    except Exception as hata:
        db.rollback()
        return templates.TemplateResponse("stok/urunler.html", urun_sayfasi_verisi(request, db, hatalar=[f"Aktarım kaydedilemedi: {hata}"]), status_code=400)
    return RedirectResponse("/urunler?aktarim=basarili", status_code=303)


@router.get("/receteler", response_class=HTMLResponse)
def receteler(request: Request, db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(STOK))):
    data = template_data(request)
    data["receteler"] = db.query(Recete).order_by(Recete.id.desc()).all()
    return templates.TemplateResponse("stok/receteler.html", data)
