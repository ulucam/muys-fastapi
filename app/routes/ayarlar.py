import io
import json
import secrets
import warnings
from pathlib import Path
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, Response, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.context import template_data
from app.database import get_db
from app.models.musteri import Musteri
from app.models.urun import Urun
from app.models.personel import Personel
from app.models.istasyon import Istasyon
from app.models.makine import Makine
from app.models.firma_ayarlari import FirmaAyarlari
from app.models.islem_logu import IslemLogu
from app.models.excel_aktarim_taslagi import ExcelAktarimTaslagi
from app.roles import ADMIN
from app.security import yetki_kontrol
from app.services.islem_log_service import islem_logla

router = APIRouter()

templates = Jinja2Templates(
    directory="app/templates"
)

ILLER_DOSYASI = Path(__file__).resolve().parent.parent / "data" / "iller.js"
MUSTERI_SUTUNLARI = ["Firma Adı", "Yetkili", "Telefon", "E-Posta", "Vergi Dairesi", "Vergi No", "İl", "İlçe", "Müşteri Türü", "Adres", "Açıklama"]
TURLER = ["Alıcı", "Tedarikçi", "Satıcı"]
URUN_SUTUNLARI = ["Ürün Kodu", "Ürün Adı", "Ürün Türü", "Birim", "Mevcut Stok", "Min. Stok", "Max. Stok", "Maliyet", "Satış Fiyatı", "Açıklama", "Durum"]
URUN_TURLERI = ["Hammadde", "Yarı Mamul", "Mamul", "Ticari Mamul"]
URUN_TIP_KODLARI = {
    "Hammadde": "Hammadde", "Yarı Mamul": "YariMamul", "YariMamul": "YariMamul",
    "Mamul": "Mamul", "Ticari Mamul": "TicariMamul", "TicariMamul": "TicariMamul",
}
URUN_TIP_ETIKETLERI = {kod: etiket for etiket, kod in URUN_TIP_KODLARI.items() if etiket not in ("YariMamul", "TicariMamul")}


def il_ilce_verisi():
    with ILLER_DOSYASI.open(encoding="utf-8") as dosya:
        return json.load(dosya)


def metin(deger):
    return str(deger or "").strip()


def sayi(deger, alan):
    if deger in (None, ""):
        return 0.0
    try:
        return float(deger)
    except (TypeError, ValueError):
        raise ValueError(f"{alan} sayısal olmalı")


def sonraki_musteri_kodu(db: Session, kullanilan_kodlar: set[str]) -> str:
    """İçe aktarım süresince tekrar etmeyen bir müşteri kodu üretir."""
    en_yuksek_numara = 0
    for kod in kullanilan_kodlar:
        if kod and kod.startswith("M") and kod[1:].isdigit():
            en_yuksek_numara = max(en_yuksek_numara, int(kod[1:]))

    sira = en_yuksek_numara + 1
    kod = f"M{sira:06}"
    while kod in kullanilan_kodlar:
        sira += 1
        kod = f"M{sira:06}"
    kullanilan_kodlar.add(kod)
    return kod


def excel_satirlarini_oku(dosya_icerigi):
    # Excel masaüstü uygulaması bazı açılır liste doğrulamalarını OpenPyXL'in
    # desteklemediği x14 uzantısına dönüştürebilir. Bu uyarı hücre verisini
    # etkilemez; yalnızca doğrulama tanımını kaldırır ve aktarımı durdurmamalıdır.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"Data Validation extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\\.worksheet\\._reader",
        )
        kitap = openpyxl.load_workbook(io.BytesIO(dosya_icerigi), data_only=True)
    if "Müşteriler" not in kitap.sheetnames:
        return [], [], ["'Müşteriler' sayfası bulunamadı."]
    sayfa = kitap["Müşteriler"]
    basliklar = [metin(hucre.value) for hucre in sayfa[1]]
    if basliklar[:len(MUSTERI_SUTUNLARI)] != MUSTERI_SUTUNLARI:
        return [], [], ["Müşteriler sayfasındaki sütun başlıkları taslakla uyuşmuyor."]

    iller = il_ilce_verisi()
    satirlar, hatalar = [], []
    for sira, satir in enumerate(sayfa.iter_rows(min_row=2, values_only=True), start=2):
        if not any(satir):
            continue
        veri = dict(zip(MUSTERI_SUTUNLARI, map(metin, satir)))
        satir_hatalari = []
        if not veri["Firma Adı"]:
            satir_hatalari.append("Firma adı zorunlu")
        if veri["Müşteri Türü"] not in TURLER:
            satir_hatalari.append("Müşteri türü Alıcı, Tedarikçi veya Satıcı olmalı")
        if veri["İl"] and veri["İl"] not in iller:
            satir_hatalari.append("Geçersiz il")
        if veri["İlçe"] and (not veri["İl"] or veri["İlçe"] not in iller.get(veri["İl"], [])):
            satir_hatalari.append("İlçenin seçilen ille eşleşmesi gerekiyor")
        if satir_hatalari:
            hatalar.append(f"Satır {sira}: {', '.join(satir_hatalari)}")
        else:
            satirlar.append(veri)
    urunler = []
    if "Stok Ürünleri" in kitap.sheetnames:
        stok_sayfasi = kitap["Stok Ürünleri"]
        stok_basliklari = [metin(hucre.value) for hucre in stok_sayfasi[1]]
        if stok_basliklari[:len(URUN_SUTUNLARI)] != URUN_SUTUNLARI:
            hatalar.append("Stok Ürünleri sayfasındaki sütun başlıkları taslakla uyuşmuyor.")
        else:
            for sira, satir in enumerate(stok_sayfasi.iter_rows(min_row=2, values_only=True), start=2):
                if not any(satir):
                    continue
                veri = dict(zip(URUN_SUTUNLARI, satir))
                satir_hatalari = []
                kod, ad = metin(veri["Ürün Kodu"]), metin(veri["Ürün Adı"])
                urun_turu, birim = URUN_TIP_KODLARI.get(metin(veri["Ürün Türü"])), metin(veri["Birim"])
                if not kod:
                    satir_hatalari.append("Ürün kodu zorunlu")
                if not ad:
                    satir_hatalari.append("Ürün adı zorunlu")
                if not urun_turu:
                    satir_hatalari.append("Ürün türü Hammadde, Yarı Mamul, Mamul veya Ticari Mamul olmalı")
                if metin(veri["Durum"]) not in ("Aktif", "Pasif"):
                    satir_hatalari.append("Durum Aktif veya Pasif olmalı")
                try:
                    urun = {
                        "kodu": kod, "adi": ad, "urun_tipi": urun_turu, "birim": birim or "Adet",
                        "mevcut_stok": sayi(veri["Mevcut Stok"], "Mevcut stok"),
                        "min_stok": sayi(veri["Min. Stok"], "Min. stok"),
                        "max_stok": sayi(veri["Max. Stok"], "Max. stok"),
                        "maliyet": sayi(veri["Maliyet"], "Maliyet"),
                        "satis_fiyati": sayi(veri["Satış Fiyatı"], "Satış fiyatı"),
                        "aciklama": metin(veri["Açıklama"]),
                        "aktif": metin(veri["Durum"]) == "Aktif",
                    }
                except ValueError as hata:
                    satir_hatalari.append(str(hata))
                    urun = None
                if satir_hatalari:
                    hatalar.append(f"Stok Ürünleri satır {sira}: {', '.join(satir_hatalari)}")
                elif urun:
                    urunler.append(urun)
    return satirlar, urunler, hatalar


@router.get("/ayarlar", response_class=HTMLResponse)
async def ayarlar(request: Request):

    return templates.TemplateResponse(
        "ayarlar/index.html",
        template_data(request)
    )


@router.get("/ayarlar/excel", response_class=HTMLResponse)
async def excel(request: Request, db: Session = Depends(get_db)):
    data = template_data(request)
    data["son_aktarim"] = (
        db.query(IslemLogu)
        .filter(IslemLogu.modul == "Excel")
        .order_by(IslemLogu.created_at.desc())
        .first()
    )
    return templates.TemplateResponse(
        "ayarlar/excel.html",
        data
    )


@router.get("/ayarlar/excel/sablon")
def excel_sablon(request: Request, db: Session = Depends(get_db)):
    iller = il_ilce_verisi()
    firma = db.query(FirmaAyarlari).first()
    mevcut_musteriler = db.query(Musteri).order_by(Musteri.id).all()
    mevcut_urunler = db.query(Urun).order_by(Urun.kodu).all()
    mevcut_personeller = db.query(Personel).order_by(Personel.kodu).all()
    mevcut_istasyonlar = db.query(Istasyon).order_by(Istasyon.kodu).all()
    mevcut_makineler = db.query(Makine).order_by(Makine.kodu).all()
    kitap = openpyxl.Workbook()
    sistem = kitap.active
    sistem.title = "Sistem Bilgileri"
    sistem.append(["MÜYS Excel Aktarım ve Dışa Aktarma"])
    sistem.append(["Açıklama", "Firma, müşteri, stok ve üretim ana verileri ayrı sayfalarda yer alır."])
    sistem.append(["Kurallar", "Zorunlu alanları doldurun; müşteri ve ürün türlerini açılır listelerden seçin."])
    sistem.append(["Dışa Aktarım Tarihi", datetime.now().strftime("%d.%m.%Y %H:%M")])
    sistem.append(["Sistem", "MÜYS v0.1.1 / FastAPI / SQLite"])
    sistem.append(["Kayıt Özeti", f"{len(mevcut_musteriler)} müşteri, {len(mevcut_urunler)} stok ürünü, {len(mevcut_personeller)} personel, {len(mevcut_istasyonlar)} istasyon, {len(mevcut_makineler)} makine"])
    sistem.append([])
    sistem.append(["Firma Bilgileri"])
    for etiket, alan in [("Firma Adı", "firma_adi"), ("Vergi No", "vergi_no"), ("Vergi Dairesi", "vergi_dairesi"), ("Telefon", "telefon"), ("E-Posta", "email"), ("Web Sitesi", "web_sitesi"), ("Adres", "adres")]:
        sistem.append([etiket, getattr(firma, alan, "") if firma else ""])
    sistem.column_dimensions["A"].width = 28
    sistem.column_dimensions["B"].width = 100
    for hucre in sistem[1]:
        hucre.font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")

    musteriler = kitap.create_sheet("Müşteriler")
    musteriler.append(MUSTERI_SUTUNLARI)
    for musteri in mevcut_musteriler:
        musteriler.append([
            musteri.firma_adi, musteri.yetkili, musteri.telefon, musteri.email,
            musteri.vergi_dairesi, musteri.vergi_no, musteri.il, musteri.ilce,
            musteri.musteri_turu or "Alıcı", musteri.adres, musteri.aciklama,
        ])
    musteriler.freeze_panes = "A2"
    son_satir = max(501, len(mevcut_musteriler) + 1)
    musteriler.auto_filter.ref = f"A1:K{son_satir}"
    for hucre in musteriler[1]:
        hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    for sutun, genislik in zip("ABCDEFGHIJK", [28, 22, 18, 28, 20, 16, 18, 20, 18, 45, 35]):
        musteriler.column_dimensions[sutun].width = genislik

    stok = kitap.create_sheet("Stok Ürünleri")
    stok.append(URUN_SUTUNLARI)
    for urun in mevcut_urunler:
        stok.append([
            urun.kodu, urun.adi, URUN_TIP_ETIKETLERI.get(urun.urun_tipi, urun.urun_tipi or "Mamul"), urun.birim or "Adet",
            urun.mevcut_stok or 0, urun.min_stok or 0, urun.max_stok or 0,
            urun.maliyet or 0, urun.satis_fiyati or 0, urun.aciklama or "",
            "Aktif" if urun.aktif else "Pasif",
        ])
    stok.freeze_panes = "A2"
    stok_son_satir = max(501, len(mevcut_urunler) + 1)
    stok.auto_filter.ref = f"A1:K{stok_son_satir}"
    for hucre in stok[1]:
        hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    for sutun, genislik in zip("ABCDEFGHIJK", [18, 32, 18, 14, 16, 16, 16, 14, 16, 36, 12]):
        stok.column_dimensions[sutun].width = genislik

    istasyon_kodlari = {istasyon.id: istasyon.kodu for istasyon in mevcut_istasyonlar}
    for sayfa_adi, basliklar, satirlar, genislikler in [
        ("Personel Listesi", ["Personel Kodu", "Ad Soyad", "Departman", "Görev", "Durum"], [[p.kodu, p.ad_soyad, p.departman, p.gorev, "Aktif" if p.aktif else "Pasif"] for p in mevcut_personeller], [18, 30, 22, 22, 14]),
        ("İstasyon Listesi", ["İstasyon Kodu", "İstasyon Adı", "Bölüm", "Açıklama", "Durum"], [[i.kodu, i.adi, i.bolum, i.aciklama, "Aktif" if i.aktif else "Pasif"] for i in mevcut_istasyonlar], [18, 30, 22, 36, 14]),
        ("Makine Listesi", ["Makine Kodu", "Makine Adı", "İstasyon Kodu", "Model", "Kapasite", "Durum"], [[m.kodu, m.adi, istasyon_kodlari.get(m.istasyon_id, ""), m.model, m.kapasite, "Aktif" if m.aktif else "Pasif"] for m in mevcut_makineler], [18, 30, 18, 22, 18, 14]),
    ]:
        sayfa = kitap.create_sheet(sayfa_adi)
        sayfa.append(basliklar)
        for satir in satirlar:
            sayfa.append(satir)
        sayfa.freeze_panes = "A2"
        sayfa.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(basliklar))}{max(2, len(satirlar) + 1)}"
        for hucre in sayfa[1]:
            hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
        for index, genislik in enumerate(genislikler, start=1):
            sayfa.column_dimensions[openpyxl.utils.get_column_letter(index)].width = genislik

    listeler = kitap.create_sheet("Listeler")
    listeler.append(["İller", "İlçeler", "Müşteri Türleri", "Ürün Türleri", "Durumlar"])
    tum_ilceler = sorted({ilce for ilceler in iller.values() for ilce in ilceler})
    durumlar = ["Aktif", "Pasif"]
    for sira in range(max(len(iller), len(tum_ilceler), len(TURLER), len(URUN_TURLERI), len(durumlar))):
        listeler.append([
            sorted(iller)[sira] if sira < len(iller) else None,
            tum_ilceler[sira] if sira < len(tum_ilceler) else None,
            TURLER[sira] if sira < len(TURLER) else None,
            URUN_TURLERI[sira] if sira < len(URUN_TURLERI) else None,
            durumlar[sira] if sira < len(durumlar) else None,
        ])
    listeler.sheet_state = "hidden"
    for formül, alan in [("'Listeler'!$A$2:$A$82", f"G2:G{son_satir}"), (f"'Listeler'!$B$2:$B${len(tum_ilceler) + 1}", f"H2:H{son_satir}"), ("'Listeler'!$C$2:$C$4", f"I2:I{son_satir}")]:
        dogrulama = DataValidation(type="list", formula1=formül, allow_blank=True)
        musteriler.add_data_validation(dogrulama)
        dogrulama.add(alan)

    for formül, alan in [
        ("'Listeler'!$D$2:$D$5", f"C2:C{stok_son_satir}"),
        ("'Listeler'!$E$2:$E$3", f"K2:K{stok_son_satir}"),
    ]:
        dogrulama = DataValidation(type="list", formula1=formül, allow_blank=False)
        stok.add_data_validation(dogrulama)
        dogrulama.add(alan)

    akis = io.BytesIO()
    kitap.save(akis)
    islem_logla(db, request, "Excel", "Excel şablonu indirildi", f"{len(mevcut_musteriler)} müşteri, {len(mevcut_urunler)} stok ürünü, {len(mevcut_personeller)} personel, {len(mevcut_istasyonlar)} istasyon ve {len(mevcut_makineler)} makine dışa aktarıldı")
    db.commit()
    return Response(
        akis.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=muys-musteri-aktarim-taslagi.xlsx"},
    )


@router.post("/ayarlar/excel/onizleme", response_class=HTMLResponse)
async def excel_onizleme(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    dosya_adi = file.filename or "adsız dosya"
    if not dosya_adi.lower().endswith(".xlsx"):
        hatalar = ["Yalnızca .xlsx uzantılı Excel dosyaları yüklenebilir."]
        satirlar, urunler = [], []
    else:
        try:
            dosya_icerigi = await file.read()
            if not dosya_icerigi:
                raise ValueError("Dosya boş")
            satirlar, urunler, hatalar = excel_satirlarini_oku(dosya_icerigi)
        except Exception as hata:
            satirlar, urunler = [], []
            hatalar = [f"Excel dosyası okunamadı: {type(hata).__name__}. Ayrıntı işlem geçmişine kaydedildi."]
            islem_logla(db, request, "Excel", "Excel önizlemesi başarısız", f"Dosya: {dosya_adi}. Hata: {type(hata).__name__}: {hata}")
            db.commit()

    mevcutlar = {m.firma_adi.casefold(): m for m in db.query(Musteri).all()}
    eklenecek, guncellenecek = [], []
    for satir in satirlar:
        (guncellenecek if satir["Firma Adı"].casefold() in mevcutlar else eklenecek).append(satir["Firma Adı"])
    if not hatalar and not satirlar and not urunler:
        hatalar.append("Aktarılacak müşteri veya stok ürünü bulunamadı.")
    if hatalar:
        islem_logla(db, request, "Excel", "Excel önizlemesi geçersiz", f"Dosya: {dosya_adi}. {len(hatalar)} doğrulama hatası bulundu.")
        db.commit()
    else:
        eski_token = request.session.pop("excel_onay_token", None)
        if eski_token:
            db.query(ExcelAktarimTaslagi).filter(ExcelAktarimTaslagi.token == eski_token).delete()
        token = secrets.token_urlsafe(32)
        db.add(ExcelAktarimTaslagi(
            token=token,
            veri=json.dumps(
                {"musteriler": satirlar, "urunler": urunler, "dosya_adi": dosya_adi},
                ensure_ascii=False,
            ),
        ))
        # SessionMiddleware istemci tarafı çerez kullanır; buraya Excel satırlarını
        # koymak çerez boyutu sınırını aşar. Sadece küçük taslak anahtarı tutulur.
        request.session["excel_onay_token"] = token
        islem_logla(db, request, "Excel", "Excel önizlemesi hazır", f"Dosya: {dosya_adi}. {len(satirlar)} müşteri, {len(urunler)} stok ürünü onay bekliyor.")
        db.commit()

    data = template_data(request)
    data["son_aktarim"] = db.query(IslemLogu).filter(IslemLogu.modul == "Excel").order_by(IslemLogu.created_at.desc()).first()
    data.update({"hatalar": hatalar, "eklenecek": eklenecek, "guncellenecek": guncellenecek, "gecerli_satir": len(satirlar), "gecerli_urun": len(urunler)})
    return templates.TemplateResponse("ayarlar/excel.html", data)


@router.post("/ayarlar/excel/onayla")
def excel_onayla(request: Request, db: Session = Depends(get_db)):
    token = request.session.get("excel_onay_token")
    taslak = db.query(ExcelAktarimTaslagi).filter(ExcelAktarimTaslagi.token == token).first() if token else None
    aktarim = json.loads(taslak.veri) if taslak else {}
    satirlar = aktarim.get("musteriler", [])
    urunler = aktarim.get("urunler", [])
    dosya_adi = aktarim.get("dosya_adi", "adsız dosya")
    if not satirlar and not urunler:
        islem_logla(db, request, "Excel", "Excel aktarımı başarısız", "Onaylanacak geçerli veri bulunamadı; önizleme süresi dolmuş veya dosya geçersiz.")
        db.commit()
        return RedirectResponse("/ayarlar/excel", status_code=303)
    try:
        kullanilan_kodlar = {
            kod for (kod,) in db.query(Musteri.musteri_kodu).all() if kod
        }
        for satir in satirlar:
            musteri = db.query(Musteri).filter(Musteri.firma_adi.ilike(satir["Firma Adı"])).first()
            if not musteri:
                musteri = Musteri(musteri_kodu=sonraki_musteri_kodu(db, kullanilan_kodlar))
                db.add(musteri)
            musteri.firma_adi = satir["Firma Adı"]
            musteri.yetkili = satir["Yetkili"]
            musteri.telefon = satir["Telefon"]
            musteri.email = satir["E-Posta"]
            musteri.vergi_dairesi = satir["Vergi Dairesi"]
            musteri.vergi_no = satir["Vergi No"]
            musteri.il, musteri.ilce = satir["İl"], satir["İlçe"]
            musteri.musteri_turu = satir["Müşteri Türü"]
            musteri.adres, musteri.aciklama = satir["Adres"], satir["Açıklama"]
        for satir in urunler:
            urun = db.query(Urun).filter(Urun.kodu == satir["kodu"]).first()
            if not urun:
                urun = Urun(kodu=satir["kodu"])
                db.add(urun)
            for alan, deger in satir.items():
                setattr(urun, alan, deger)
        db.delete(taslak)
        request.session.pop("excel_onay_token", None)
        islem_logla(db, request, "Excel", "Excel aktarımı tamamlandı", f"Dosya: {dosya_adi}. {len(satirlar)} müşteri ve {len(urunler)} stok ürünü aktarıldı/güncellendi.")
        db.commit()
    except Exception as hata:
        db.rollback()
        islem_logla(db, request, "Excel", "Excel aktarımı başarısız", f"Dosya: {dosya_adi}. Hata: {type(hata).__name__}: {hata}")
        db.commit()
    return RedirectResponse("/ayarlar/excel", status_code=303)


@router.get("/ayarlar/yedek", response_class=HTMLResponse)
async def yedek(request: Request):

    return templates.TemplateResponse(
        "ayarlar/yedek.html",
        template_data(request)
    )


@router.get("/ayarlar/loglar", response_class=HTMLResponse)
async def loglar(
    request: Request,
    db: Session = Depends(get_db),
    yetki=Depends(yetki_kontrol(ADMIN)),
):
    data = template_data(request)
    data["loglar"] = db.query(IslemLogu).order_by(IslemLogu.created_at.desc()).limit(500).all()

    return templates.TemplateResponse("ayarlar/loglar.html", data)


@router.get("/ayarlar/firma", response_class=HTMLResponse)
async def firma(request: Request, db: Session = Depends(get_db)):
    data = template_data(request)
    data["firma"] = db.query(FirmaAyarlari).first()

    return templates.TemplateResponse("ayarlar/firma.html", data)


@router.post("/ayarlar/firma")
async def firma_kaydet(
    request: Request,
    firma_adi: str = Form(""), vergi_no: str = Form(""),
    vergi_dairesi: str = Form(""), telefon: str = Form(""),
    email: str = Form(""), web_sitesi: str = Form(""), adres: str = Form(""),
    db: Session = Depends(get_db),
):
    firma = db.query(FirmaAyarlari).first()
    if not firma:
        firma = FirmaAyarlari()
        db.add(firma)
    firma.firma_adi, firma.vergi_no = firma_adi, vergi_no
    firma.vergi_dairesi, firma.telefon = vergi_dairesi, telefon
    firma.email, firma.web_sitesi, firma.adres = email, web_sitesi, adres
    islem_logla(db, request, "Ayarlar", "Firma bilgileri güncellendi", firma_adi)
    db.commit()

    return RedirectResponse("/ayarlar/firma", status_code=303)


@router.get("/ayarlar/sistem", response_class=HTMLResponse)
async def sistem(request: Request):

    return templates.TemplateResponse(
        "ayarlar/sistem.html",
        template_data(request)
    )
