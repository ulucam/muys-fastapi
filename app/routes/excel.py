from fastapi import APIRouter, Depends, Response, HTTPException
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
from app.database import get_db
from app.models import Musteri, Urun, Siparis, User
from app.auth import get_current_user, rol_gerekli

router = APIRouter(prefix="/api/excel", tags=["excel"])

@router.get("/musteriler")
@rol_gerekli(['Admin', 'Muhasebe'])
def excel_musteriler(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Musteriler"
    
    # Başlıklar
    basliklar = ['ID', 'Firma Kodu', 'Firma Adı', 'Yetkili', 'Telefon', 'Email', 'İl', 'İlçe', 'Tip']
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=baslik)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Veriler
    musteriler = db.query(Musteri).all()
    for row_idx, m in enumerate(musteriler, 2):
        ws.cell(row=row_idx, column=1, value=m.id)
        ws.cell(row=row_idx, column=2, value=m.firma_kodu)
        ws.cell(row=row_idx, column=3, value=m.firma_adi)
        ws.cell(row=row_idx, column=4, value=m.yetkili or '')
        ws.cell(row=row_idx, column=5, value=m.telefon or '')
        ws.cell(row=row_idx, column=6, value=m.email or '')
        ws.cell(row=row_idx, column=7, value=m.il or '')
        ws.cell(row=row_idx, column=8, value=m.ilce or '')
        ws.cell(row=row_idx, column=9, value=m.musteri_tipi or '')
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=musteriler.xlsx"}
    )

@router.get("/siparisler")
@rol_gerekli(['Admin', 'Muhasebe', 'Uretim'])
def excel_siparisler(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Siparisler"
    
    basliklar = ['Sipariş No', 'Müşteri', 'Tarih', 'Teslim Tarihi', 'Durum', 'Notlar']
    for col, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col, value=baslik)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    siparisler = db.query(Siparis).order_by(Siparis.created_at.desc()).all()
    for row_idx, s in enumerate(siparisler, 2):
        ws.cell(row=row_idx, column=1, value=s.siparis_no)
        ws.cell(row=row_idx, column=2, value=s.musteri.firma_adi if s.musteri else '')
        ws.cell(row=row_idx, column=3, value=s.siparis_tarihi.strftime('%d.%m.%Y') if s.siparis_tarihi else '')
        ws.cell(row=row_idx, column=4, value=s.teslim_tarihi.strftime('%d.%m.%Y') if s.teslim_tarihi else '')
        ws.cell(row=row_idx, column=5, value=s.durum)
        ws.cell(row=row_idx, column=6, value=s.notlar or '')
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=siparisler.xlsx"}
    )
