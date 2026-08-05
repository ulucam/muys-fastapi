import io
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import HTMLResponse, Response, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.context import template_data
from app.database import get_db
from app.models.istasyon import Istasyon
from app.models.makine import Makine
from app.models.personel import Personel
from app.models.personel_makine import PersonelMakine
from app.models.puantaj import Puantaj
from app.models.recete import Recete
from app.models.recete_kalem import ReceteKalem
from app.models.urun import Urun
from app.models.urun_sinif_operasyon import UrunSinifOperasyon
from app.models.urun_sinifi import UrunSinifi
from app.roles import YONETIM
from app.security import yetki_kontrol
from app.services.islem_log_service import islem_logla

router = APIRouter(tags=["Üretim Tanımları"])
templates = Jinja2Templates(directory="app/templates")

SAYFALAR = {
    "Personeller": ["Personel Kodu", "Ad Soyad", "Departman", "Görev", "Durum"],
    "İstasyonlar": ["İstasyon Kodu", "İstasyon Adı", "Bölüm", "Açıklama", "Durum"],
    "Makineler": ["Makine Kodu", "Makine Adı", "İstasyon Kodu", "Model", "Kapasite", "Durum"],
    "Personel Makine Atamaları": ["Personel Kodu", "Makine Kodu", "Rol", "Hedef Performans", "Durum"],
    "Ürün Sınıfları": ["Sınıf Kodu", "Sınıf Adı", "Açıklama", "Durum"],
    "Sınıf Reçete Operasyonları": ["Sınıf Kodu", "Sıra", "İstasyon Kodu", "Makine Kodu", "Operasyon Adı", "Hedef Çevrim Süresi (dk)", "Kontrol Noktası", "Durum"],
    "Ürünler": ["Ürün Kodu", "Ürün Adı", "Ürün Türü", "Ürün Sınıfı Kodu", "Birim", "Mevcut Stok", "Min. Stok", "Max. Stok", "Maliyet", "Satış Fiyatı", "Açıklama", "Durum"],
    "Ürün Reçetesi": ["Üst Ürün Kodu", "Bileşen Ürün Kodu", "Miktar", "Birim", "Fire Oranı (%)", "Sıra"],
}
URUN_TIPLERI = {"Hammadde", "YariMamul", "Mamul", "TicariMamul"}
AKTIF_DURUMLAR = {"Aktif": True, "Pasif": False}


def metin(deger):
    return str(deger or "").strip()


def sayi(deger, alan, tam_sayi=False):
    try:
        sonuc = int(deger) if tam_sayi else float(deger)
    except (TypeError, ValueError):
        raise ValueError(f"{alan} sayısal olmalı")
    return sonuc


def sayfa_verisi(kitap, sayfa_adi):
    if sayfa_adi not in kitap.sheetnames:
        raise ValueError(f"'{sayfa_adi}' sayfası bulunamadı")
    sayfa = kitap[sayfa_adi]
    basliklar = [metin(h.value) for h in sayfa[1]]
    if basliklar[:len(SAYFALAR[sayfa_adi])] != SAYFALAR[sayfa_adi]:
        raise ValueError(f"'{sayfa_adi}' sütunları şablonla uyuşmuyor")
    return [dict(zip(SAYFALAR[sayfa_adi], satir)) for satir in sayfa.iter_rows(min_row=2, values_only=True) if any(satir)]


def durum(deger):
    sonuc = AKTIF_DURUMLAR.get(metin(deger))
    if sonuc is None:
        raise ValueError("Durum Aktif veya Pasif olmalı")
    return sonuc


def sayfa_ekle(kitap, ad, satirlar):
    sayfa = kitap.create_sheet(ad)
    sayfa.append(SAYFALAR[ad])
    for satir in satirlar:
        sayfa.append(satir)
    sayfa.freeze_panes = "A2"
    sayfa.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(SAYFALAR[ad]))}{max(2, len(satirlar) + 1)}"
    for hucre in sayfa[1]:
        hucre.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        hucre.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    for sutun in range(1, len(SAYFALAR[ad]) + 1):
        sayfa.column_dimensions[openpyxl.utils.get_column_letter(sutun)].width = 22
    return sayfa


def ekran_verisi(request, db, **ek):
    atamalar = db.query(PersonelMakine).filter(PersonelMakine.aktif.is_(True)).all()
    makine_haritasi = {m.id: m for m in db.query(Makine).all()}
    personel_atamalari = {}
    for atama in atamalar:
        personel_atamalari.setdefault(atama.personel_id, []).append((atama, makine_haritasi.get(atama.makine_id)))
    personel_puantajlari = {}
    for puantaj in db.query(Puantaj).order_by(Puantaj.tarih.desc()).limit(500).all():
        if len(personel_puantajlari.setdefault(puantaj.personel_id, [])) < 10:
            personel_puantajlari[puantaj.personel_id].append(puantaj)
    data = template_data(request)
    data.update({
        "personel_sayisi": db.query(Personel).filter(Personel.aktif.is_(True)).count(),
        "istasyon_sayisi": db.query(Istasyon).filter(Istasyon.aktif.is_(True)).count(),
        "makine_sayisi": db.query(Makine).filter(Makine.aktif.is_(True)).count(),
        "urun_sinifi_sayisi": db.query(UrunSinifi).filter(UrunSinifi.aktif.is_(True)).count(),
        "istasyonlar": db.query(Istasyon).order_by(Istasyon.kodu).all(),
        "personeller": db.query(Personel).order_by(Personel.kodu).all(),
        "makineler": db.query(Makine).order_by(Makine.kodu).all(),
        "urun_siniflari": db.query(UrunSinifi).order_by(UrunSinifi.kodu).all(),
        "urunler": db.query(Urun).order_by(Urun.kodu).all(),
        "personel_atamalari": personel_atamalari,
        "personel_puantajlari": personel_puantajlari,
    })
    data.update(ek)
    return data


@router.get("/uretim-tanimlari", response_class=HTMLResponse)
def uretim_tanimlari(request: Request, error: str | None = None, goster: str = "", db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(YONETIM))):
    hata = "Kayıt kaydedilemedi. Zorunlu alanları ve seçilen ilişkileri kontrol edin." if error else None
    listeler = {
        "personeller": ("Aktif Personeller", db.query(Personel).filter(Personel.aktif.is_(True)).order_by(Personel.kodu).all()),
        "istasyonlar": ("Aktif İstasyonlar", db.query(Istasyon).filter(Istasyon.aktif.is_(True)).order_by(Istasyon.kodu).all()),
        "makineler": ("Aktif Makineler", db.query(Makine).filter(Makine.aktif.is_(True)).order_by(Makine.kodu).all()),
        "urun_siniflari": ("Aktif Ürün Sınıfları", db.query(UrunSinifi).filter(UrunSinifi.aktif.is_(True)).order_by(UrunSinifi.kodu).all()),
    }
    baslik, kayitlar = listeler.get(goster, (None, []))
    return templates.TemplateResponse("uretim/tanimlar.html", ekran_verisi(request, db, hata=hata, goster=goster, liste_basligi=baslik, secili_kayitlar=kayitlar))


@router.get("/uretim-tanimlari/duzenle/{tip}/{kod}", response_class=HTMLResponse)
def duzenle_form(tip: str, kod: str, request: Request, db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(YONETIM))):
    modeller = {"personel": Personel, "istasyon": Istasyon, "makine": Makine, "sinif": UrunSinifi}
    model = modeller.get(tip)
    kayit = db.query(model).filter(model.kodu == kod).first() if model else None
    if not kayit:
        return RedirectResponse("/uretim-tanimlari", status_code=303)
    data = ekran_verisi(request, db)
    data.update({"duzenle_tipi": tip, "kayit": kayit})
    return templates.TemplateResponse("uretim/duzenle.html", data)


@router.post("/uretim-tanimlari/kaydet/{tip}")
async def manuel_kaydet(tip: str, request: Request, db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(YONETIM))):
    form = await request.form()
    aktif = form.get("aktif") == "true"
    try:
        if tip == "personel":
            kod = metin(form.get("kodu"))
            nesne = db.query(Personel).filter(Personel.kodu == kod).first() or Personel(kodu=kod)
            if not kod or not metin(form.get("ad_soyad")):
                raise ValueError("Personel kodu ve ad soyad zorunlu")
            nesne.ad_soyad, nesne.departman, nesne.gorev, nesne.aktif = metin(form.get("ad_soyad")), metin(form.get("departman")), metin(form.get("gorev")), aktif
        elif tip == "istasyon":
            kod = metin(form.get("kodu"))
            nesne = db.query(Istasyon).filter(Istasyon.kodu == kod).first() or Istasyon(kodu=kod)
            if not kod or not metin(form.get("adi")):
                raise ValueError("İstasyon kodu ve adı zorunlu")
            nesne.adi, nesne.bolum, nesne.aciklama, nesne.aktif = metin(form.get("adi")), metin(form.get("bolum")), metin(form.get("aciklama")), aktif
        elif tip == "makine":
            kod, istasyon_kodu = metin(form.get("kodu")), metin(form.get("istasyon_kodu"))
            istasyon = db.query(Istasyon).filter(Istasyon.kodu == istasyon_kodu).first()
            if not kod or not metin(form.get("adi")) or not istasyon:
                raise ValueError("Makine kodu, adı ve istasyon seçimi zorunlu")
            nesne = db.query(Makine).filter(Makine.kodu == kod).first() or Makine(kodu=kod)
            nesne.adi, nesne.istasyon_id, nesne.model, nesne.kapasite, nesne.aktif = metin(form.get("adi")), istasyon.id, metin(form.get("model")), metin(form.get("kapasite")), aktif
        elif tip == "atama":
            personel = db.query(Personel).filter(Personel.kodu == metin(form.get("personel_kodu"))).first()
            makine = db.query(Makine).filter(Makine.kodu == metin(form.get("makine_kodu"))).first()
            if not personel or not makine:
                raise ValueError("Personel ve makine seçimi zorunlu")
            nesne = db.query(PersonelMakine).filter(PersonelMakine.personel_id == personel.id, PersonelMakine.makine_id == makine.id).first() or PersonelMakine(personel_id=personel.id, makine_id=makine.id)
            nesne.rol, nesne.hedef_performans, nesne.aktif = metin(form.get("rol")) or "Operatör", sayi(form.get("hedef_performans") or 100, "Hedef performans"), aktif
        elif tip == "sinif":
            kod = metin(form.get("kodu"))
            nesne = db.query(UrunSinifi).filter(UrunSinifi.kodu == kod).first() or UrunSinifi(kodu=kod)
            if not kod or not metin(form.get("adi")):
                raise ValueError("Sınıf kodu ve adı zorunlu")
            nesne.adi, nesne.aciklama, nesne.aktif = metin(form.get("adi")), metin(form.get("aciklama")), aktif
        elif tip == "operasyon":
            sinif = db.query(UrunSinifi).filter(UrunSinifi.kodu == metin(form.get("sinif_kodu"))).first()
            istasyon = db.query(Istasyon).filter(Istasyon.kodu == metin(form.get("istasyon_kodu"))).first()
            makine_kodu = metin(form.get("makine_kodu"))
            makine = db.query(Makine).filter(Makine.kodu == makine_kodu).first() if makine_kodu else None
            sira = sayi(form.get("sira"), "Sıra", tam_sayi=True)
            if not sinif or not istasyon or not metin(form.get("operasyon_adi")):
                raise ValueError("Ürün sınıfı, istasyon, sıra ve operasyon adı zorunlu")
            if makine and makine.istasyon_id != istasyon.id:
                raise ValueError("Makine seçilen istasyona bağlı olmalı")
            nesne = db.query(UrunSinifOperasyon).filter(UrunSinifOperasyon.urun_sinifi_id == sinif.id, UrunSinifOperasyon.sira_no == sira).first() or UrunSinifOperasyon(urun_sinifi_id=sinif.id, sira_no=sira)
            nesne.istasyon_id, nesne.makine_id, nesne.operasyon_adi = istasyon.id, makine.id if makine else None, metin(form.get("operasyon_adi"))
            nesne.hedef_cevrim_suresi, nesne.kontrol_noktasi, nesne.aktif = sayi(form.get("hedef_cevrim") or 0, "Hedef çevrim"), metin(form.get("kontrol_noktasi")), aktif
        elif tip == "urun":
            kod, sinif_kodu, urun_tipi = metin(form.get("kodu")), metin(form.get("sinif_kodu")), metin(form.get("urun_tipi"))
            sinif = db.query(UrunSinifi).filter(UrunSinifi.kodu == sinif_kodu).first() if sinif_kodu else None
            if not kod or not metin(form.get("adi")) or urun_tipi not in URUN_TIPLERI or (sinif_kodu and not sinif):
                raise ValueError("Ürün kodu, adı, türü ve varsa ürün sınıfı seçimi geçerli olmalı")
            nesne = db.query(Urun).filter(Urun.kodu == kod).first() or Urun(kodu=kod)
            nesne.adi, nesne.urun_tipi, nesne.urun_sinifi_id = metin(form.get("adi")), urun_tipi, sinif.id if sinif else None
            nesne.birim, nesne.mevcut_stok, nesne.min_stok = metin(form.get("birim")) or "Adet", sayi(form.get("mevcut_stok") or 0, "Mevcut stok"), sayi(form.get("min_stok") or 0, "Min. stok")
            nesne.aktif = aktif
        elif tip == "recete":
            ust = db.query(Urun).filter(Urun.kodu == metin(form.get("ust_urun_kodu"))).first()
            bilesen = db.query(Urun).filter(Urun.kodu == metin(form.get("bilesen_urun_kodu"))).first()
            if not ust or not bilesen or ust.id == bilesen.id:
                raise ValueError("Geçerli ve birbirinden farklı üst ürün ile bileşen ürün seçin")
            recete = db.query(Recete).filter(Recete.urun_id == ust.id).first()
            if not recete:
                recete = Recete(urun_id=ust.id, recete_no=f"R-{ust.kodu}", aciklama=f"{ust.adi} reçetesi")
                db.add(recete)
                db.flush()
            nesne = db.query(ReceteKalem).filter(ReceteKalem.recete_id == recete.id, ReceteKalem.malzeme_id == bilesen.id).first() or ReceteKalem(recete_id=recete.id, malzeme_id=bilesen.id)
            nesne.miktar, nesne.birim, nesne.fire_orani, nesne.sira_no, nesne.aktif = sayi(form.get("miktar"), "Miktar"), metin(form.get("birim")) or bilesen.birim, sayi(form.get("fire_orani") or 0, "Fire oranı"), sayi(form.get("sira") or 1, "Sıra", tam_sayi=True), aktif
        else:
            raise ValueError("Bilinmeyen kayıt türü")
        if nesne not in db:
            db.add(nesne)
        islem_logla(db, request, "Üretim", "Üretim tanımı kaydedildi", tip)
        db.commit()
        return RedirectResponse("/uretim-tanimlari", status_code=303)
    except Exception:
        db.rollback()
        return RedirectResponse("/uretim-tanimlari?error=kayit", status_code=303)


@router.get("/uretim-tanimlari/excel-sablon")
def excel_sablon(request: Request, db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(YONETIM))):
    kitap = openpyxl.Workbook()
    bilgi = kitap.active
    bilgi.title = "Sistem Bilgileri"
    bilgi.append(["MÜYS Üretim Ana Veri Aktarımı"])
    bilgi.append(["Açıklama", "Sayfaları sırayla doldurun: istasyonlar → makineler → personel/makine atamaları → ürün sınıfları → operasyonlar."])
    bilgi.append(["Tarih", datetime.now().strftime("%d.%m.%Y %H:%M")])
    bilgi.column_dimensions["A"].width = 28
    bilgi.column_dimensions["B"].width = 115

    sayfa_ekle(kitap, "Personeller", [[p.kodu, p.ad_soyad, p.departman, p.gorev, "Aktif" if p.aktif else "Pasif"] for p in db.query(Personel).order_by(Personel.kodu)])
    sayfa_ekle(kitap, "İstasyonlar", [[i.kodu, i.adi, i.bolum, i.aciklama, "Aktif" if i.aktif else "Pasif"] for i in db.query(Istasyon).order_by(Istasyon.kodu)])
    istasyon_kodlari = {i.id: i.kodu for i in db.query(Istasyon).all()}
    sayfa_ekle(kitap, "Makineler", [[m.kodu, m.adi, istasyon_kodlari.get(m.istasyon_id, ""), m.model, m.kapasite, "Aktif" if m.aktif else "Pasif"] for m in db.query(Makine).order_by(Makine.kodu)])
    personel_kodlari = {p.id: p.kodu for p in db.query(Personel).all()}
    makine_kodlari = {m.id: m.kodu for m in db.query(Makine).all()}
    sayfa_ekle(kitap, "Personel Makine Atamaları", [[personel_kodlari.get(a.personel_id, ""), makine_kodlari.get(a.makine_id, ""), a.rol, a.hedef_performans, "Aktif" if a.aktif else "Pasif"] for a in db.query(PersonelMakine).all()])
    sayfa_ekle(kitap, "Ürün Sınıfları", [[s.kodu, s.adi, s.aciklama, "Aktif" if s.aktif else "Pasif"] for s in db.query(UrunSinifi).order_by(UrunSinifi.kodu)])
    sinif_kodlari = {s.id: s.kodu for s in db.query(UrunSinifi).all()}
    sayfa_ekle(kitap, "Sınıf Reçete Operasyonları", [[sinif_kodlari.get(o.urun_sinifi_id, ""), o.sira_no, istasyon_kodlari.get(o.istasyon_id, ""), makine_kodlari.get(o.makine_id, ""), o.operasyon_adi, o.hedef_cevrim_suresi, o.kontrol_noktasi, "Aktif" if o.aktif else "Pasif"] for o in db.query(UrunSinifOperasyon).order_by(UrunSinifOperasyon.urun_sinifi_id, UrunSinifOperasyon.sira_no)])
    sayfa_ekle(kitap, "Ürünler", [[u.kodu, u.adi, u.urun_tipi, sinif_kodlari.get(u.urun_sinifi_id, ""), u.birim, u.mevcut_stok, u.min_stok, u.max_stok, u.maliyet, u.satis_fiyati, u.aciklama, "Aktif" if u.aktif else "Pasif"] for u in db.query(Urun).order_by(Urun.kodu)])
    urun_kodlari = {u.id: u.kodu for u in db.query(Urun).all()}
    recete_kalemleri = db.query(ReceteKalem, Recete).join(Recete, ReceteKalem.recete_id == Recete.id).all()
    sayfa_ekle(kitap, "Ürün Reçetesi", [[urun_kodlari.get(r.urun_id, ""), urun_kodlari.get(k.malzeme_id, ""), k.miktar, k.birim, k.fire_orani, k.sira_no] for k, r in recete_kalemleri])

    listeler = kitap.create_sheet("Listeler")
    listeler.append(["Durumlar", "Ürün Türleri"])
    for sira in range(4):
        listeler.append([["Aktif", "Pasif"][sira] if sira < 2 else None, ["Hammadde", "YariMamul", "Mamul", "TicariMamul"][sira]])
    listeler.sheet_state = "hidden"
    for ad, sutun in [("Personeller", "E"), ("İstasyonlar", "E"), ("Makineler", "F"), ("Personel Makine Atamaları", "E"), ("Ürün Sınıfları", "D"), ("Sınıf Reçete Operasyonları", "H"), ("Ürünler", "L")]:
        dogrulama = DataValidation(type="list", formula1="'Listeler'!$A$2:$A$3")
        kitap[ad].add_data_validation(dogrulama)
        dogrulama.add(f"{sutun}2:{sutun}1000")
    dogrulama = DataValidation(type="list", formula1="'Listeler'!$B$2:$B$5")
    kitap["Ürünler"].add_data_validation(dogrulama)
    dogrulama.add("C2:C1000")

    akis = io.BytesIO()
    kitap.save(akis)
    islem_logla(db, request, "Üretim", "Üretim Excel şablonu indirildi", "Personel, istasyon, makine, ürün sınıfı ve reçete şablonu")
    db.commit()
    return Response(akis.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=muys-uretim-ana-veri.xlsx"})


@router.post("/uretim-tanimlari/excel-aktar", response_class=HTMLResponse)
async def excel_aktar(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), yetki=Depends(yetki_kontrol(YONETIM))):
    dosya_adi = file.filename or "adsız dosya"
    try:
        if not dosya_adi.lower().endswith(".xlsx"):
            raise ValueError("Yalnızca .xlsx dosyası yüklenebilir")
        kitap = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
        veriler = {ad: sayfa_verisi(kitap, ad) for ad in SAYFALAR}

        personeller = {x.kodu: x for x in db.query(Personel).all()}
        for satir in veriler["Personeller"]:
            kod = metin(satir["Personel Kodu"])
            if not kod or not metin(satir["Ad Soyad"]):
                raise ValueError("Personeller: Personel Kodu ve Ad Soyad zorunlu")
            nesne = personeller.get(kod) or Personel(kodu=kod)
            nesne.ad_soyad, nesne.departman, nesne.gorev = metin(satir["Ad Soyad"]), metin(satir["Departman"]), metin(satir["Görev"])
            nesne.aktif = durum(satir["Durum"])
            if nesne not in db:
                db.add(nesne)
            personeller[kod] = nesne
        db.flush()

        istasyonlar = {x.kodu: x for x in db.query(Istasyon).all()}
        for satir in veriler["İstasyonlar"]:
            kod = metin(satir["İstasyon Kodu"])
            if not kod or not metin(satir["İstasyon Adı"]):
                raise ValueError("İstasyonlar: İstasyon Kodu ve İstasyon Adı zorunlu")
            nesne = istasyonlar.get(kod) or Istasyon(kodu=kod)
            nesne.adi, nesne.bolum, nesne.aciklama = metin(satir["İstasyon Adı"]), metin(satir["Bölüm"]), metin(satir["Açıklama"])
            nesne.aktif = durum(satir["Durum"])
            if nesne not in db:
                db.add(nesne)
            istasyonlar[kod] = nesne
        db.flush()

        makineler = {x.kodu: x for x in db.query(Makine).all()}
        for satir in veriler["Makineler"]:
            kod, istasyon_kodu = metin(satir["Makine Kodu"]), metin(satir["İstasyon Kodu"])
            if not kod or not metin(satir["Makine Adı"]) or istasyon_kodu not in istasyonlar:
                raise ValueError("Makineler: Makine Kodu, Makine Adı ve geçerli İstasyon Kodu zorunlu")
            nesne = makineler.get(kod) or Makine(kodu=kod)
            nesne.adi, nesne.istasyon_id = metin(satir["Makine Adı"]), istasyonlar[istasyon_kodu].id
            nesne.model, nesne.kapasite, nesne.aktif = metin(satir["Model"]), metin(satir["Kapasite"]), durum(satir["Durum"])
            if nesne not in db:
                db.add(nesne)
            makineler[kod] = nesne
        db.flush()

        siniflar = {x.kodu: x for x in db.query(UrunSinifi).all()}
        for satir in veriler["Ürün Sınıfları"]:
            kod = metin(satir["Sınıf Kodu"])
            if not kod or not metin(satir["Sınıf Adı"]):
                raise ValueError("Ürün Sınıfları: Sınıf Kodu ve Sınıf Adı zorunlu")
            nesne = siniflar.get(kod) or UrunSinifi(kodu=kod)
            nesne.adi, nesne.aciklama, nesne.aktif = metin(satir["Sınıf Adı"]), metin(satir["Açıklama"]), durum(satir["Durum"])
            if nesne not in db:
                db.add(nesne)
            siniflar[kod] = nesne
        db.flush()

        urunler = {x.kodu: x for x in db.query(Urun).all()}
        for satir in veriler["Ürünler"]:
            kod, sinif_kodu, urun_tipi = metin(satir["Ürün Kodu"]), metin(satir["Ürün Sınıfı Kodu"]), metin(satir["Ürün Türü"])
            if not kod or not metin(satir["Ürün Adı"]) or urun_tipi not in URUN_TIPLERI:
                raise ValueError("Ürünler: Ürün Kodu, Ürün Adı ve geçerli Ürün Türü zorunlu")
            if sinif_kodu and sinif_kodu not in siniflar:
                raise ValueError(f"Ürünler: '{sinif_kodu}' ürün sınıfı bulunamadı")
            nesne = urunler.get(kod) or Urun(kodu=kod)
            nesne.adi, nesne.urun_tipi, nesne.urun_sinifi_id = metin(satir["Ürün Adı"]), urun_tipi, siniflar[sinif_kodu].id if sinif_kodu else None
            nesne.birim = metin(satir["Birim"]) or "Adet"
            nesne.mevcut_stok = sayi(satir["Mevcut Stok"] or 0, "Mevcut Stok")
            nesne.min_stok = sayi(satir["Min. Stok"] or 0, "Min. Stok")
            nesne.max_stok = sayi(satir["Max. Stok"] or 0, "Max. Stok")
            nesne.maliyet = sayi(satir["Maliyet"] or 0, "Maliyet")
            nesne.satis_fiyati = sayi(satir["Satış Fiyatı"] or 0, "Satış Fiyatı")
            nesne.aciklama, nesne.aktif = metin(satir["Açıklama"]), durum(satir["Durum"])
            if nesne not in db:
                db.add(nesne)
            urunler[kod] = nesne
        db.flush()

        for satir in veriler["Personel Makine Atamaları"]:
            personel, makine = personeller.get(metin(satir["Personel Kodu"])), makineler.get(metin(satir["Makine Kodu"]))
            if not personel or not makine:
                raise ValueError("Personel Makine Atamaları: geçerli Personel Kodu ve Makine Kodu zorunlu")
            nesne = db.query(PersonelMakine).filter(PersonelMakine.personel_id == personel.id, PersonelMakine.makine_id == makine.id).first()
            if not nesne:
                nesne = PersonelMakine(personel_id=personel.id, makine_id=makine.id)
                db.add(nesne)
            nesne.rol, nesne.hedef_performans, nesne.aktif = metin(satir["Rol"]) or "Operatör", sayi(satir["Hedef Performans"] or 100, "Hedef Performans"), durum(satir["Durum"])

        for satir in veriler["Sınıf Reçete Operasyonları"]:
            sinif, istasyon = siniflar.get(metin(satir["Sınıf Kodu"])), istasyonlar.get(metin(satir["İstasyon Kodu"]))
            makine_kodu = metin(satir["Makine Kodu"])
            makine = makineler.get(makine_kodu) if makine_kodu else None
            if not sinif or not istasyon or not metin(satir["Operasyon Adı"]):
                raise ValueError("Sınıf Reçete Operasyonları: sınıf, istasyon ve operasyon adı zorunlu")
            if makine and makine.istasyon_id != istasyon.id:
                raise ValueError("Sınıf Reçete Operasyonları: makine seçilen istasyona bağlı olmalı")
            sira = sayi(satir["Sıra"], "Sıra", tam_sayi=True)
            nesne = db.query(UrunSinifOperasyon).filter(UrunSinifOperasyon.urun_sinifi_id == sinif.id, UrunSinifOperasyon.sira_no == sira).first()
            if not nesne:
                nesne = UrunSinifOperasyon(urun_sinifi_id=sinif.id, sira_no=sira)
                db.add(nesne)
            nesne.istasyon_id, nesne.makine_id = istasyon.id, makine.id if makine else None
            nesne.operasyon_adi = metin(satir["Operasyon Adı"])
            nesne.hedef_cevrim_suresi = sayi(satir["Hedef Çevrim Süresi (dk)"] or 0, "Hedef Çevrim Süresi")
            nesne.kontrol_noktasi, nesne.aktif = metin(satir["Kontrol Noktası"]), durum(satir["Durum"])

        for satir in veriler["Ürün Reçetesi"]:
            ust, bilesen = urunler.get(metin(satir["Üst Ürün Kodu"])), urunler.get(metin(satir["Bileşen Ürün Kodu"]))
            if not ust or not bilesen:
                raise ValueError("Ürün Reçetesi: geçerli üst ürün ve bileşen ürün kodu zorunlu")
            recete = db.query(Recete).filter(Recete.urun_id == ust.id).first()
            if not recete:
                recete = Recete(urun_id=ust.id, recete_no=f"R-{ust.kodu}", aciklama=f"{ust.adi} reçetesi")
                db.add(recete)
                db.flush()
            kalem = db.query(ReceteKalem).filter(ReceteKalem.recete_id == recete.id, ReceteKalem.malzeme_id == bilesen.id).first()
            if not kalem:
                kalem = ReceteKalem(recete_id=recete.id, malzeme_id=bilesen.id)
                db.add(kalem)
            kalem.miktar, kalem.birim = sayi(satir["Miktar"], "Miktar"), metin(satir["Birim"]) or bilesen.birim
            kalem.fire_orani, kalem.sira_no, kalem.aktif = sayi(satir["Fire Oranı (%)"] or 0, "Fire Oranı"), sayi(satir["Sıra"], "Sıra", tam_sayi=True), True

        toplam = sum(len(satirlar) for satirlar in veriler.values())
        islem_logla(db, request, "Üretim", "Üretim ana verisi aktarıldı", f"Dosya: {dosya_adi}. {toplam} satır işlendi.")
        db.commit()
        return templates.TemplateResponse("uretim/tanimlar.html", ekran_verisi(request, db, basari=f"{toplam} satır başarıyla işlendi."))
    except Exception as hata:
        db.rollback()
        islem_logla(db, request, "Üretim", "Üretim Excel aktarımı başarısız", f"Dosya: {dosya_adi}. {type(hata).__name__}: {hata}")
        db.commit()
        return templates.TemplateResponse("uretim/tanimlar.html", ekran_verisi(request, db, hata=str(hata)))
